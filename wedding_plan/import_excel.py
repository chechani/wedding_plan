"""
Multi-sheet Excel import — the mechanism that makes "define everything as a
new marriage" real (see erpnext-whatsapp-plan.md section B6, which this
follows: master data first, batches traceable, reconcile before you rely on
the numbers). One workbook, one sheet per doctype, sheets processed in a
sensible default order — though a link whose target isn't in the DB yet will
pull that target's own sheet in out of order (see `links` below), so exact
sheet ordering in the workbook isn't load-bearing.

Each SHEETS entry describes:
  doctype       - the DocType to upsert into
  sheet_name    - the .xlsx sheet tab name (also the row 1 header row is
                   read literally as fieldnames, so a downloadable template
                   generated from this same table always matches)
  natural_key   - fieldnames that together identify "the same row" across
                   repeated imports (upsert key)
  links         - {fieldname: (link_doctype, lookup_fieldname)} — the sheet
                   holds a human-readable value (e.g. a venue name); this
                   resolves it to the linked doc's name *within this
                   wedding*. If it's not in the DB yet, the linked doctype's
                   own sheet (if present in this workbook) is searched for a
                   matching row, and that row is imported first — out of
                   SHEETS order, recursively, so a chain of unresolved links
                   still resolves. Only a value that appears nowhere — not
                   in the DB and not as a row in its own sheet — is a
                   row-level error (matching B6's "imports fail on missing
                   links"): nothing is ever fabricated, so a typo in a venue
                   name still can't silently spawn a duplicate venue, it can
                   only be satisfied by data the user actually provided.
  multiselects  - {fieldname: (link_doctype, lookup_fieldname)} for
                   comma-separated text columns that map onto a Table
                   MultiSelect field (e.g. Guests' "Invited To" column).
"""
import json

import frappe
from openpyxl import Workbook, load_workbook

SHEETS = [
    {"doctype": "WD Sub Group", "sheet_name": "SubGroups", "natural_key": ["sub_group_name"]},
    {"doctype": "WD Venue", "sheet_name": "Venues", "natural_key": ["venue_name"]},
    {"doctype": "WD Function", "sheet_name": "Functions", "natural_key": ["function_name"],
     "links": {"venue": ("WD Venue", "venue_name")}},
    {"doctype": "WD Room", "sheet_name": "Rooms", "natural_key": ["venue", "room_no"],
     "links": {"venue": ("WD Venue", "venue_name")}},
    {"doctype": "WD Team", "sheet_name": "Teams", "natural_key": ["team_key"]},
    {"doctype": "WD Vendor", "sheet_name": "Vendors", "natural_key": ["vendor_name"]},
    {"doctype": "WD Guest", "sheet_name": "Guests", "natural_key": ["household_name", "primary_phone"],
     "links": {"sub_group": ("WD Sub Group", "sub_group_name"), "venue": ("WD Venue", "venue_name")},
     "multiselects": {"functions_invited": ("WD Function", "function_name")}},
    {"doctype": "WD Room Allotment", "sheet_name": "RoomAllotments", "natural_key": ["room", "guest"],
     "links": {"room": ("WD Room", "room_no"), "guest": ("WD Guest", "household_name")}},
    {"doctype": "WD Run Sheet Item", "sheet_name": "RunSheet", "natural_key": ["date", "time", "title"],
     "links": {"function": ("WD Function", "function_name"), "venue": ("WD Venue", "venue_name")}},
    {"doctype": "WD Anchor Brief", "sheet_name": "AnchorBriefs", "natural_key": ["function"],
     "links": {"function": ("WD Function", "function_name")}},
    {"doctype": "WD Tech Requirement", "sheet_name": "TechRequirements", "natural_key": ["function"],
     "links": {"function": ("WD Function", "function_name")}},
    {"doctype": "WD Meal Session", "sheet_name": "MealSessions", "natural_key": ["date", "session_name", "venue"],
     "links": {"venue": ("WD Venue", "venue_name"), "vendor": ("WD Vendor", "vendor_name")}},
    {"doctype": "WD Vehicle", "sheet_name": "Vehicles", "natural_key": ["vehicle_number"],
     "links": {"vendor": ("WD Vendor", "vendor_name")}},
    {"doctype": "WD Convoy Leg", "sheet_name": "ConvoyLegs", "natural_key": ["leg_no", "date"],
     "links": {"from_venue": ("WD Venue", "venue_name"), "to_venue": ("WD Venue", "venue_name")}},
    {"doctype": "WD Pickup", "sheet_name": "Pickups", "natural_key": ["guest", "eta"],
     "links": {"guest": ("WD Guest", "household_name"), "vehicle": ("WD Vehicle", "vehicle_number")}},
    {"doctype": "WD Crate", "sheet_name": "Crates", "natural_key": ["crate_no"],
     "links": {"destination_venue": ("WD Venue", "venue_name")}},
    {"doctype": "WD Shagun Entry", "sheet_name": "ShagunEntries", "natural_key": ["guest", "function", "received_by"],
     "links": {"guest": ("WD Guest", "household_name"), "function": ("WD Function", "function_name")}},
    {"doctype": "WD Approval Matrix Entry", "sheet_name": "ApprovalMatrix", "natural_key": ["situation"]},
    {"doctype": "WD Risk", "sheet_name": "Risks", "natural_key": ["title"]},
    {"doctype": "WD Gap Note", "sheet_name": "GapNotes", "natural_key": ["title"]},
]


def _sheet_fieldnames(doctype):
    # "wedding" is always injected server-side from the import call's target
    # wedding, never read from the sheet — a stray value in that column
    # would silently reassign a row to the wrong tenant.
    meta = frappe.get_meta(doctype)
    fields = [f for f in meta.fields if f.fieldname != "wedding"]
    return [f.fieldname for f in fields if f.fieldtype != "Table MultiSelect"] + \
        [f.fieldname for f in fields if f.fieldtype == "Table MultiSelect"]


def build_template_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in SHEETS:
        ws = wb.create_sheet(sheet["sheet_name"])
        ws.append(_sheet_fieldnames(sheet["doctype"]))
    return wb


class _ImportContext:
    """Per-run state: the workbook, a doctype->sheet index for pull-ins,
    a cache of each sheet's non-blank rows (shared between the main loop
    and out-of-order pulls so a sheet is only read off disk once), the
    resolved-link cache, and bookkeeping to make pulled-in rows idempotent
    and cycles detectable."""

    def __init__(self, wedding, wb):
        self.wedding = wedding
        self.wb = wb
        self.sheet_cfg_by_doctype = {s["doctype"]: s for s in SHEETS}
        self._sheet_rows = {}
        self.link_cache = {}
        self.imported_keys = set()  # (doctype, natural_key_tuple) already saved this run
        self._pulling = set()  # (doctype, normalized lookup value) — cycle guard

    def rows_for_sheet(self, sheet_name):
        if sheet_name not in self._sheet_rows:
            raw_rows = list(self.wb[sheet_name].iter_rows(values_only=True))
            parsed = []
            if raw_rows:
                headers = [str(h).strip() for h in raw_rows[0] if h is not None]
                for row_idx, raw_row in enumerate(raw_rows[1:], start=2):
                    row = dict(zip(headers, raw_row))
                    if any(v not in (None, "") for v in row.values()):
                        parsed.append((row_idx, row))
            self._sheet_rows[sheet_name] = parsed
        return self._sheet_rows[sheet_name]


def _resolve_link(ctx, doctype, lookup_field, value):
    if not value:
        return None, None
    key = (doctype, str(value).strip().lower())
    if key in ctx.link_cache:
        return ctx.link_cache[key], None

    name = frappe.db.get_value(doctype, {"wedding": ctx.wedding, lookup_field: value}, "name")
    if not name:
        name = _pull_in_linked_row(ctx, doctype, lookup_field, key)
    if not name:
        return None, f"{doctype} '{value}' not found — import {doctype} sheet first"

    ctx.link_cache[key] = name
    return name, None


def _pull_in_linked_row(ctx, doctype, lookup_field, key):
    """Look for a row matching `key` in `doctype`'s own sheet elsewhere in
    the workbook and import it now, so a link isn't limited to targets that
    happen to appear earlier in SHEETS order. Returns None (never raises)
    when there's nothing to pull in, so the caller falls through to the
    normal not-found error; raises only if the pulled-in row itself fails
    to import (e.g. one of *its* links is genuinely missing)."""
    source_sheet = ctx.sheet_cfg_by_doctype.get(doctype)
    if not source_sheet or source_sheet["sheet_name"] not in ctx.wb.sheetnames:
        return None
    if key in ctx._pulling:
        return None  # circular reference between sheets — fall through to not-found

    match = None
    for _, candidate in ctx.rows_for_sheet(source_sheet["sheet_name"]):
        cell = candidate.get(lookup_field)
        if cell not in (None, "") and str(cell).strip().lower() == key[1]:
            match = candidate
            break
    if match is None:
        return None

    ctx._pulling.add(key)
    try:
        return _import_row(ctx, source_sheet, match)
    except Exception as e:
        raise ValueError(
            f"while auto-importing {doctype} '{key[1]}' from the {source_sheet['sheet_name']} sheet: {e}"
        ) from e
    finally:
        ctx._pulling.discard(key)


def run_import(wedding: str, file_path: str, import_job: str = None):
    wb = load_workbook(file_path, data_only=True)
    ctx = _ImportContext(wedding, wb)
    totals = {"rows_total": 0, "rows_succeeded": 0, "rows_failed": 0}
    errors = []
    sheets_processed = []

    for sheet_cfg in SHEETS:
        sheet_name = sheet_cfg["sheet_name"]
        if sheet_name not in wb.sheetnames:
            continue
        rows = ctx.rows_for_sheet(sheet_name)
        if not rows:
            continue
        sheet_result = {"sheet": sheet_name, "doctype": sheet_cfg["doctype"], "succeeded": 0, "failed": 0}

        for row_idx, row in rows:
            natural_key = tuple(row.get(f) for f in sheet_cfg["natural_key"])
            if (sheet_cfg["doctype"], natural_key) in ctx.imported_keys:
                continue  # a later sheet already pulled this row in out of order

            totals["rows_total"] += 1
            try:
                _import_row(ctx, sheet_cfg, row)
                totals["rows_succeeded"] += 1
                sheet_result["succeeded"] += 1
            except Exception as e:  # noqa: BLE001 — one bad row must not abort the batch
                totals["rows_failed"] += 1
                sheet_result["failed"] += 1
                errors.append({"sheet": sheet_name, "row": row_idx, "error": str(e)})

        sheets_processed.append(sheet_result)

    if import_job:
        frappe.db.set_value("WD Import Job", import_job, {
            "status": "Done" if totals["rows_failed"] == 0 else "Done",
            "rows_total": totals["rows_total"],
            "rows_succeeded": totals["rows_succeeded"],
            "rows_failed": totals["rows_failed"],
            "error_log": json.dumps(errors),
        })
        frappe.db.commit()

    return {**totals, "sheets_processed": sheets_processed, "errors": errors}


def _import_row(ctx, sheet_cfg, row):
    doctype = sheet_cfg["doctype"]
    links = sheet_cfg.get("links", {})
    multiselects = sheet_cfg.get("multiselects", {})

    data = {"wedding": ctx.wedding}
    for field, value in row.items():
        if field in ("wedding", None) or field in links or field in multiselects or value in (None, ""):
            continue
        data[field] = value

    for field, (link_doctype, lookup_field) in links.items():
        raw_value = row.get(field)
        resolved, err = _resolve_link(ctx, link_doctype, lookup_field, raw_value)
        if err:
            raise ValueError(err)
        if resolved is not None:
            data[field] = resolved

    multiselect_rows = {}
    for field, (link_doctype, lookup_field) in multiselects.items():
        raw_value = row.get(field)
        if not raw_value:
            continue
        names = []
        for part in str(raw_value).split(","):
            part = part.strip()
            if not part:
                continue
            resolved, err = _resolve_link(ctx, link_doctype, lookup_field, part)
            if err:
                raise ValueError(err)
            names.append(resolved)
        multiselect_rows[field] = names

    natural_key = sheet_cfg["natural_key"]
    filters = {"wedding": ctx.wedding}
    for key_field in natural_key:
        filters[key_field] = data.get(key_field)

    existing_name = frappe.db.get_value(doctype, filters, "name")
    if existing_name:
        doc = frappe.get_doc(doctype, existing_name)
        doc.update(data)
    else:
        doc = frappe.new_doc(doctype)
        doc.update(data)

    for field, names in multiselect_rows.items():
        link_fieldname_in_child = _multiselect_child_link_field(doctype, field)
        doc.set(field, [])
        for name in names:
            doc.append(field, {link_fieldname_in_child: name})

    doc.save(ignore_permissions=True)
    ctx.imported_keys.add((doctype, tuple(row.get(f) for f in natural_key)))
    return doc.name


def _multiselect_child_link_field(doctype, fieldname):
    meta = frappe.get_meta(doctype)
    df = meta.get_field(fieldname)
    child_meta = frappe.get_meta(df.options)
    link_fields = [f.fieldname for f in child_meta.fields if f.fieldtype == "Link"]
    return link_fields[0]
